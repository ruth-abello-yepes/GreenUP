"""Exportacion personal sin acceder a Supabase ni crear datos reales."""
import base64
from datetime import datetime, timedelta, timezone
from io import BytesIO
import os
from pathlib import Path
import sys
import unittest
from unittest.mock import patch

import jwt
from flask import Flask
from openpyxl import load_workbook

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))
from app.controllers.reportes_routes import reportes_bp


class ReportesCiudadanoTest(unittest.TestCase):
    def setUp(self):
        self.secreto = "greenup-test-reportes-no-produccion-2026"
        self.env = patch.dict(os.environ, {"JWT_SECRET_KEY": self.secreto})
        self.env.start()
        app = Flask(__name__)
        app.register_blueprint(reportes_bp)
        app.logger.disabled = True
        self.cliente = app.test_client()
        self.consulta = patch("app.services.reportes_service.listar_reporte_reciclaje")
        self.modelo = self.consulta.start()
        self.modelo.return_value = [self.fila(2, 12.5), self.fila(1, 7), self.fila(3, 3)]
        self.perfiles = patch("app.services.reporte_ciudadano_service.obtener_perfil_usuario", return_value={"usuario": "laura.suarez", "nombres": "Laura Sofia", "apellidos": "Suarez Perez"})
        self.perfil = self.perfiles.start()

    def tearDown(self):
        self.consulta.stop()
        self.perfiles.stop()
        self.env.stop()

    @staticmethod
    def fila(estado=2, cantidad=12.5):
        return {"fecha_hora": datetime(2026, 9, 3, 10), "id_estado": estado, "cantidad": cantidad,
                "material": "Plástico", "punto": "Centro ecológico", "estado": ""}

    def cabeceras(self, rol=3, expirado=False):
        token = jwt.encode({"id_usuario": 42, "id_rol": rol,
                            "exp": datetime.now(timezone.utc) + timedelta(minutes=-1 if expirado else 5)},
                           self.secreto, algorithm="HS256")
        return {"Authorization": f"Bearer {token}"}

    def get(self, query="", **kwargs):
        return self.cliente.get("/api/reportes/ciudadano" + query, headers=self.cabeceras(**kwargs))

    def test_exige_sesion_y_rol_ciudadano(self):
        self.assertEqual(self.cliente.get("/api/reportes/ciudadano").status_code, 401)
        self.assertEqual(self.get(expirado=True).status_code, 401)
        for rol in (1, 2):
            self.assertEqual(self.get(rol=rol).status_code, 403)
        self.modelo.assert_not_called()

    def test_usuario_de_url_no_puede_cambiar_propietario(self):
        r = self.get("?id_usuario=999&fecha_inicio=2026-09-01&fecha_fin=2026-09-03")
        self.assertEqual(r.status_code, 200)
        self.modelo.assert_called_once_with({"id_usuario": 42, "fecha_inicio": "2026-09-01", "fecha_fin": "2026-09-03"})
        self.assertEqual(r.json["kg_confirmados"], 12.5)
        self.assertEqual(r.json["total_registros"], 3)
        self.assertNotIn("usuario", r.json["registros"][0])
        self.assertIn("no-store", r.headers["Cache-Control"])

    def test_rechaza_filtros_invalidos_antes_de_consultar(self):
        for query in ("?fecha_inicio=2026-02-30", "?fecha_inicio=2026-09-10&fecha_fin=2026-09-01", "?formato=html", "?formato=vista&tipo=html"):
            self.assertEqual(self.get(query).status_code, 400)
        self.modelo.assert_not_called()

    def test_pdf_identifica_usuario_real_y_fecha_colombiana_incluso_sin_registros(self):
        self.modelo.return_value = []
        instante = datetime(2026, 9, 3, 16, 25, 40, tzinfo=timezone(timedelta(hours=-5)))
        with patch("app.services.reporte_ciudadano_service.datetime", wraps=datetime) as reloj:
            reloj.now.return_value = instante
            r = self.get("?formato=vista&tipo=pdf&id_usuario=999&usuario=otro&generado_en=2000-01-01")
        self.assertEqual(r.status_code, 200)
        self.perfil.assert_called_once_with(42)
        self.assertEqual(r.json["usuario"], "laura.suarez")
        self.assertEqual(r.json["nombre_mostrar"], "Laura Suarez")
        self.assertEqual(r.json["generado_en"], "2026-09-03T16:25:40-05:00")
        self.assertTrue(base64.b64decode(r.json["archivo"]["base64"]).startswith(b"%PDF-"))

    def test_excel_conserva_filas_numeros_y_texto_sin_ejecutar_formulas(self):
        self.modelo.return_value = [dict(self.fila(), material="=1+1", punto="=HYPERLINK(\"https://example.com\")")] * 125
        r = self.get("?formato=vista&tipo=excel")
        self.assertEqual(r.status_code, 200)
        libro = load_workbook(BytesIO(base64.b64decode(r.json["archivo"]["base64"])))
        self.assertEqual(libro["Registros"].max_row, 126)
        self.assertEqual(libro["Registros"]["B2"].data_type, "s")
        self.assertEqual(libro["Registros"]["D2"].data_type, "s")
        self.assertEqual(libro["Resumen"]["A8"].data_type, "s")
        self.assertEqual(libro["Registros"]["C2"].value, 12.5)
        self.assertEqual(libro["Resumen"]["B4"].value, r.json["kg_confirmados"])
        self.assertEqual(len(libro["Resumen"]._charts), 2)
        self.modelo.assert_called_once()

    def test_pdf_multipagina_y_vista_corresponden_a_una_consulta(self):
        self.modelo.return_value = [dict(self.fila(), punto="Punto ecológico de reciclaje " * 4)] * 125
        r = self.get("?formato=vista&tipo=pdf")
        self.assertEqual(r.status_code, 200)
        contenido = base64.b64decode(r.json["archivo"]["base64"])
        self.assertTrue(contenido.startswith(b"%PDF-"))
        self.assertGreater(contenido.count(b"/Type /Page\n"), 1)
        self.assertEqual(len(r.json["registros"]), 125)
        self.modelo.assert_called_once()

    def test_reportes_vacios_se_pueden_descargar(self):
        self.modelo.return_value = []
        for formato, mime in (("pdf", "application/pdf"), ("excel", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")):
            r = self.get(f"?formato={formato}")
            self.assertEqual(r.status_code, 200)
            self.assertEqual(r.mimetype, mime)
            self.assertIn("attachment", r.headers["Content-Disposition"])

    def test_error_de_base_de_datos_no_expone_detalles(self):
        self.modelo.side_effect = RuntimeError("credencial-privada")
        r = self.get()
        self.assertEqual(r.status_code, 503)
        self.assertNotIn("credencial-privada", r.get_data(as_text=True))


if __name__ == "__main__":
    unittest.main()
