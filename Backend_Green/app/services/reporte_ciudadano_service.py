"""Reporte personal: una misma seleccion alimenta la vista previa, PDF y Excel."""
from collections import Counter, defaultdict
from io import BytesIO
from xml.sax.saxutils import escape

from app.services.reportes_service import servicio_reporte_reciclaje


def preparar_reporte_ciudadano(id_usuario, argumentos):
    filtros = {campo: argumentos.get(campo) for campo in ("fecha_inicio", "fecha_fin")}
    # El propietario procede exclusivamente del JWT, nunca de la URL.
    filtros["id_usuario"] = id_usuario
    datos, estado = servicio_reporte_reciclaje(filtros)
    if estado != 200:
        return datos, estado
    filas, materiales, estados = [], defaultdict(float), Counter()
    for item in datos:
        texto_estado = str(item.get("estado") or "").lower()
        confirmado = texto_estado == "confirmado" or item.get("id_estado") == 2
        estado_fila = "Confirmado" if confirmado else (
            "Rechazado" if texto_estado == "rechazado" or item.get("id_estado") == 3 else "Pendiente"
        )
        cantidad = float(item.get("cantidad") or 0)
        material = item.get("material") or "Sin clasificar"
        fecha = item.get("fecha_hora")
        filas.append({
            "fecha": fecha.date().isoformat() if fecha else "",
            "material": material, "cantidad": cantidad,
            "punto": item.get("punto") or "Sin punto", "estado": estado_fila,
        })
        estados[estado_fila] += 1
        if confirmado:
            materiales[material] += cantidad
    desglose = sorted(materiales.items(), key=lambda item: (-item[1], item[0]))
    # Mantiene legible el grafico; la tabla conserva todos los registros.
    grafico = desglose[:10]
    if len(desglose) > 10:
        grafico.append(("Otros materiales", sum(valor for _, valor in desglose[10:])))
    return {
        "fecha_inicio": filtros["fecha_inicio"], "fecha_fin": filtros["fecha_fin"],
        "periodo": f"{filtros['fecha_inicio'] or 'Desde el primer registro'} / {filtros['fecha_fin'] or 'Hasta hoy'}",
        "total_registros": len(filas), "kg_confirmados": round(sum(materiales.values()), 2),
        "estados": [{"estado": nombre, "cantidad": estados[nombre]} for nombre in ("Confirmado", "Pendiente", "Rechazado")],
        "materiales": [{"material": nombre, "cantidad": round(valor, 2)} for nombre, valor in grafico],
        "registros": filas,
    }, 200


COLUMNAS = ["Fecha", "Material", "Cantidad (kg)", "Punto de reciclaje", "Estado"]


def generar_pdf_ciudadano(reporte):
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.pagesizes import A4
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
    from reportlab.graphics.shapes import Drawing
    from reportlab.graphics.charts.barcharts import HorizontalBarChart

    salida = BytesIO()
    estilos = getSampleStyleSheet()
    estilos.add(ParagraphStyle(name="CeldaGreenUp", fontName="Helvetica", fontSize=8, leading=11, wordWrap="CJK"))
    celda = lambda valor: Paragraph(escape(str(valor)), estilos["CeldaGreenUp"])
    contenido = [Paragraph("GreenUp | Mi reporte de reciclaje", estilos["Title"]),
                 Paragraph(escape(reporte["periodo"]), estilos["Normal"]), Spacer(1, 14),
                 Paragraph(f"{reporte['total_registros']} registros | {reporte['kg_confirmados']:g} kg confirmados", estilos["Heading2"]),
                 Paragraph("Los kilogramos y el grafico incluyen solo entregas confirmadas. La tabla incluye todos los estados.", estilos["Normal"]), Spacer(1, 14)]
    if reporte["materiales"]:
        altura = max(130, 26 * len(reporte["materiales"]))
        dibujo = Drawing(490, altura + 35)
        grafica = HorizontalBarChart()
        grafica.x, grafica.y, grafica.width, grafica.height = 145, 25, 325, altura
        grafica.data = [[item["cantidad"] for item in reporte["materiales"]]]
        grafica.categoryAxis.categoryNames = [item["material"][:25] for item in reporte["materiales"]]
        grafica.categoryAxis.labels.fontSize = 8
        grafica.valueAxis.valueMin = 0
        grafica.bars[0].fillColor = colors.HexColor("#296c1f")
        dibujo.add(grafica)
        contenido.extend([Paragraph("Material confirmado (kg)", estilos["Heading3"]), dibujo, Spacer(1, 14)])
    estados = " | ".join(f"{item['estado']}: {item['cantidad']}" for item in reporte["estados"])
    contenido.extend([Paragraph(estados, estilos["Normal"]), Spacer(1, 14)])
    filas = [COLUMNAS]
    for fila in reporte["registros"]:
        filas.append([celda(fila["fecha"]), celda(fila["material"]), celda(f"{fila['cantidad']:g}"), celda(fila["punto"]), celda(fila["estado"])])
    if not reporte["registros"]:
        contenido.append(Paragraph("No hay registros en el periodo seleccionado.", estilos["Normal"]))
    tabla = Table(filas, colWidths=[65, 110, 80, 155, 85], repeatRows=1)
    tabla.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#003d6c")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 8), ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f0f7ed")]),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 8), ("TOPPADDING", (0, 0), (-1, -1), 8),
        ("LINEBELOW", (0, 0), (-1, 0), 1, colors.HexColor("#296c1f")),
    ]))
    contenido.append(tabla)

    def pie(canvas, documento):
        canvas.saveState()
        canvas.setFont("Helvetica", 8)
        canvas.drawString(50, 25, "GreenUp - Reporte personal")
        canvas.drawRightString(A4[0] - 50, 25, f"Pagina {documento.page}")
        canvas.restoreState()

    SimpleDocTemplate(salida, pagesize=A4, rightMargin=50, leftMargin=50,
                      topMargin=40, bottomMargin=45, title="Mi reporte de reciclaje - GreenUp").build(
                          contenido, onFirstPage=pie, onLaterPages=pie)
    salida.seek(0)
    return salida


def generar_excel_ciudadano(reporte):
    from datetime import date
    from openpyxl import Workbook
    from openpyxl.chart import BarChart, PieChart, Reference
    from openpyxl.styles import Font, PatternFill, Alignment

    libro = Workbook()
    resumen = libro.active
    resumen.title = "Resumen"
    resumen.append(["GreenUp - Mi reporte de reciclaje"])
    resumen.append(["Periodo", reporte["periodo"]])
    resumen.append(["Registros", reporte["total_registros"]])
    resumen.append(["Kg confirmados", reporte["kg_confirmados"]])
    resumen.append(["Solo las entregas confirmadas suman kilogramos."])
    resumen.append([])
    resumen.append(["Material confirmado", "Cantidad (kg)", None, "Estado", "Registros"])
    for item in reporte["materiales"]:
        resumen.append([item["material"], item["cantidad"]])
        resumen.cell(resumen.max_row, 1).data_type = "s"
    for indice, item in enumerate(reporte["estados"], 8):
        resumen.cell(indice, 4, item["estado"])
        resumen.cell(indice, 5, item["cantidad"])
    if reporte["materiales"]:
        barras = BarChart()
        barras.title, barras.y_axis.title = "Material confirmado", "kg"
        barras.add_data(Reference(resumen, min_col=2, min_row=7, max_row=7 + len(reporte["materiales"])), titles_from_data=True)
        barras.set_categories(Reference(resumen, min_col=1, min_row=8, max_row=7 + len(reporte["materiales"])))
        resumen.add_chart(barras, "G2")
    if reporte["total_registros"]:
        pastel = PieChart()
        pastel.title = "Estado de las entregas"
        pastel.add_data(Reference(resumen, min_col=5, min_row=7, max_row=10), titles_from_data=True)
        pastel.set_categories(Reference(resumen, min_col=4, min_row=8, max_row=10))
        resumen.add_chart(pastel, "G18")
    registros = libro.create_sheet("Registros")
    registros.append(COLUMNAS)
    for item in reporte["registros"]:
        registros.append([date.fromisoformat(item["fecha"]) if item["fecha"] else "", item["material"], item["cantidad"], item["punto"], item["estado"]])
        registros.cell(registros.max_row, 1).number_format = "yyyy-mm-dd"
        registros.cell(registros.max_row, 3).number_format = "0.00"
        for columna in (2, 4, 5):
            # Los nombres introducidos por usuarios siempre son texto, no formulas.
            registros.cell(registros.max_row, columna).data_type = "s"
    registros.auto_filter.ref = registros.dimensions
    registros.freeze_panes = "A2"
    for fila in registros.iter_rows(min_row=2):
        lineas = max((len(str(celda.value or "")) + ancho - 1) // ancho
                     for celda, ancho in zip(fila, (28, 28, 17, 40, 18)))
        registros.row_dimensions[fila[0].row].height = max(22, 15 * lineas)
    for hoja, encabezados in ((resumen, (1, 7)), (registros, (1,))):
        for fila in hoja:
            for celda in fila:
                celda.alignment = Alignment(vertical="top", wrap_text=True)
                if celda.row in encabezados:
                    celda.fill = PatternFill("solid", fgColor="003D6C")
                    celda.font = Font(color="FFFFFF", bold=True)
        for columna, ancho in zip("ABCDE", (30, 30, 19, 42, 20)):
            hoja.column_dimensions[columna].width = ancho
    resumen.row_dimensions[1].height = 32
    resumen.column_dimensions["B"].width = 42
    salida = BytesIO()
    libro.save(salida)
    salida.seek(0)
    return salida
